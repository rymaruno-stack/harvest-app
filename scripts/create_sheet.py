#!/usr/bin/env python3
"""
前月シートを複製して新月シートを作成する。
日付（A列）・曜日（B列）を更新し、数値データ（C〜AY列）をクリアする。

Usage:
    python scripts/create_sheet.py --template "2025出荷実績.xlsx" --year 2025 --month 12

    # 別ファイルに保存したい場合:
    python scripts/create_sheet.py --template "2025出荷実績.xlsx" --year 2026 --month 1 --output "2026出荷実績.xlsx"

依存パッケージ:
    pip install openpyxl
"""

import argparse
import sys
from calendar import monthrange
from datetime import date

import openpyxl

# 月曜=0 ... 日曜=6 の順（Python の date.weekday() に対応）
DOW_JA = ['月', '火', '水', '木', '金', '土', '日']

# ── セクション定義 ───────────────────────────────────────────────────
SECTIONS = [
    {'name': '全体',   'start_row': 6},
    {'name': '初号機', 'start_row': 46},
    {'name': '弐号機', 'start_row': 86},
    {'name': '参号機', 'start_row': 126},
]

# データ列範囲: C列(3) 〜 AY列(51)
DATA_COL_START = 3   # C
DATA_COL_END   = 51  # AY


def get_prev_month(year: int, month: int):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def clear_or_keep(cell) -> None:
    """数式セルは残し、数値・文字列・None はクリア（None に設定）する。"""
    val = cell.value
    if isinstance(val, str) and val.startswith('='):
        return  # 数式は保持
    cell.value = None


def main() -> None:
    parser = argparse.ArgumentParser(description='前月シートを複製して新月シートを作成')
    parser.add_argument('--template', required=True, help='Excel ファイルパス')
    parser.add_argument('--year',     type=int, required=True, help='新シートの年 (例: 2025)')
    parser.add_argument('--month',    type=int, required=True, help='新シートの月 (例: 12)')
    parser.add_argument('--output',   default=None,
                        help='出力ファイルパス (省略時は --template ファイルに上書き保存)')
    args = parser.parse_args()

    new_sheet_name  = f'{args.month}月'
    prev_y, prev_m  = get_prev_month(args.year, args.month)
    prev_sheet_name = f'{prev_m}月'

    wb = openpyxl.load_workbook(args.template)

    if prev_sheet_name not in wb.sheetnames:
        sys.exit(
            f'エラー: 前月シート "{prev_sheet_name}" が見つかりません。\n'
            f'利用可能なシート: {wb.sheetnames}'
        )

    if new_sheet_name in wb.sheetnames:
        print(f'シート "{new_sheet_name}" は既に存在します。処理をスキップします。')
        return

    # 前月シートを複製
    ws_src = wb[prev_sheet_name]
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = new_sheet_name

    # 新月シートをファイル末尾に移動
    wb.move_sheet(ws_new, offset=len(wb.sheetnames))
    print(f'シート "{prev_sheet_name}" → "{new_sheet_name}" に複製しました')

    _, last_day = monthrange(args.year, args.month)

    for section in SECTIONS:
        start_row = section['start_row']
        for day in range(1, 32):
            row = start_row + day - 1
            if day <= last_day:
                d = date(args.year, args.month, day)
                # A列: 日（数字）, B列: 曜日
                ws_new.cell(row=row, column=1).value = day
                ws_new.cell(row=row, column=2).value = DOW_JA[d.weekday()]
                # C〜AY列: 数値をクリア（数式は保持）
                for col in range(DATA_COL_START, DATA_COL_END + 1):
                    clear_or_keep(ws_new.cell(row=row, column=col))
            else:
                # 当月に存在しない日（例: 2月の29〜31日）は行全体をクリア
                for col in range(1, DATA_COL_END + 1):
                    ws_new.cell(row=row, column=col).value = None

    print(
        f'日付・曜日を {args.year}年{args.month}月 に更新し、'
        f'数値データをクリアしました（月の日数: {last_day}日）'
    )

    output = args.output or args.template
    wb.save(output)
    print(f'保存完了: {output}')


if __name__ == '__main__':
    main()
