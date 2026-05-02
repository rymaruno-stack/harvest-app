#!/usr/bin/env python3
"""
Supabase から月次出荷データを取得して Excel に書き込む。

Usage:
    python scripts/export_excel.py --year 2025 --month 11 \
        --template "2025出荷実績.xlsx" --output "2025出荷実績_updated.xlsx"

依存パッケージ:
    pip install openpyxl requests
"""

import argparse
import os
import sys
from calendar import monthrange
from pathlib import Path

import openpyxl
import requests

# ── 列マッピング (Excel列 → DBフィールド) ─────────────────────────────
COLUMN_MAP = [
    ('C',  'ja_nagabako_am'),
    ('D',  'ja_fukabako_am'),
    ('E',  'ja_hirabako_as'),
    ('F',  'ja_hirabako_am'),
    ('G',  'ja_10khira_am'),
    ('H',  'ja_regular_al'),
    ('I',  'ja_regular_am'),
    ('J',  'ja_regular_as'),
    ('K',  'ja_regular_a2s'),
    ('L',  'ja_regular_bl'),
    ('M',  'ja_regular_bm'),
    ('N',  'ja_regular_bs'),
    ('O',  'ja_kobukuro_al'),
    ('P',  'ja_kobukuro_bl'),
    ('Q',  'ja_kobukuro_bm'),
    ('R',  'ja_kobukuro_cm'),
    ('S',  'ja_kikakugai_dm'),
    ('T',  'ja_kikakugai_ds'),
    ('U',  'market_fukabako_a'),
    ('V',  'market_regular_as'),
    ('W',  'market_regular_am'),
    ('X',  'market_regular_al'),
    ('Y',  'market_regular_bs'),
    ('Z',  'market_regular_bm'),
    ('AA', 'market_regular_bl'),
    ('AB', 'market_10k_pori'),
    ('AC', 'jfp_contena'),
    ('AD', 'jfp_fukabako_5kg'),
    ('AE', 'jfp_b_5kg'),
    ('AF', 'jfp_cd_10kg'),
    ('AG', 'ja_uriage_nagabako'),
    ('AH', 'ja_uriage_fukabako'),
    ('AI', 'ja_uriage_hirabako'),
    ('AJ', 'ja_uriage_10k'),
    ('AK', 'ja_uriage_regular'),
    ('AL', 'ja_uriage_kobukuro'),
    ('AM', 'ja_uriage_kikakugai'),
    ('AN', 'market_uriage_fukabako_a'),
    ('AO', 'market_uriage_as'),
    ('AP', 'market_uriage_am'),
    ('AQ', 'market_uriage_al'),
    ('AR', 'market_uriage_bs'),
    ('AS', 'market_uriage_bm'),
    ('AT', 'market_uriage_bl'),
    ('AU', 'market_uriage_pori'),
    ('AV', 'jfp_uriage_contena'),
    ('AW', 'jfp_uriage_fukabako'),
    ('AX', 'jfp_uriage_b'),
    ('AY', 'jfp_uriage_cd'),
]

# ── セクション定義 ───────────────────────────────────────────────────
# 全体セクション（行6〜36）は各ハウスの合算数式が入っているため書き込み対象外
SECTIONS = [
    {'name': '初号機', 'start_row': 46,  'house_id': 1},
    {'name': '弐号機', 'start_row': 86,  'house_id': 2},
    {'name': '参号機', 'start_row': 126, 'house_id': 3},
]


def load_env(env_path: str = '.env') -> None:
    """シンプルな .env パーサー（python-dotenv 不要）"""
    path = Path(env_path)
    if not path.exists():
        return
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, _, val = line.partition('=')
                val = val.strip().strip('"').strip("'")
                os.environ.setdefault(key.strip(), val)


def fetch_records(url: str, key: str, year: int, month: int) -> list:
    """Supabase PostgREST API で harvests を取得する（supabase パッケージ不使用）。"""
    _, last_day = monthrange(year, month)
    from_date = f'{year}-{month:02d}-01'
    to_date   = f'{year}-{month:02d}-{last_day:02d}'

    endpoint = f'{url.rstrip("/")}/rest/v1/harvests'
    headers  = {
        'apikey':        key,
        'Authorization': f'Bearer {key}',
    }
    # requests では同名パラメータをタプルのリストで渡す
    params = [
        ('select', '*'),
        ('date',   f'gte.{from_date}'),
        ('date',   f'lte.{to_date}'),
        ('limit',  '200'),   # 31日×3ハウス=最大93件
        ('order',  'date.asc'),
    ]
    resp = requests.get(endpoint, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def build_daily_totals(records: list, house_id) -> dict:
    """日付 → {field: 合計値} の辞書を返す。house_id=None は全ハウス合算。"""
    totals = {}
    for rec in records:
        if house_id is not None and rec['house_id'] != house_id:
            continue
        d = rec['date']  # 'YYYY-MM-DD'
        if d not in totals:
            totals[d] = {}
        for _, field in COLUMN_MAP:
            totals[d][field] = totals[d].get(field, 0) + (rec.get(field) or 0)
    return totals


def write_section(ws, section: dict, daily: dict, year: int, month: int) -> None:
    _, last_day = monthrange(year, month)
    start_row = section['start_row']
    for day in range(1, last_day + 1):
        date_str = f'{year}-{month:02d}-{day:02d}'
        if date_str not in daily:
            continue  # Supabaseにデータなし → 元のExcel値をそのまま残す
        row = start_row + day - 1
        day_data = daily[date_str]
        for col_letter, field in COLUMN_MAP:
            ws[f'{col_letter}{row}'] = day_data.get(field, 0)


def main() -> None:
    parser = argparse.ArgumentParser(description='Supabase → Excel エクスポート')
    parser.add_argument('--year',     type=int, required=True,        help='対象年 (例: 2025)')
    parser.add_argument('--month',    type=int, required=True,        help='対象月 (例: 11)')
    parser.add_argument('--template', required=True,                  help='テンプレート Excel ファイルパス')
    parser.add_argument('--output',   required=True,                  help='出力ファイルパス')
    parser.add_argument('--env',      default='.env',                 help='.env ファイルパス (デフォルト: .env)')
    args = parser.parse_args()

    load_env(args.env)
    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_KEY') or os.environ.get('SUPABASE_ANON_KEY')
    if not url or not key:
        sys.exit(
            'エラー: SUPABASE_URL / SUPABASE_KEY が未設定です。\n'
            '.env ファイルに以下を記述してください:\n'
            '  SUPABASE_URL=https://xxxx.supabase.co\n'
            '  SUPABASE_KEY=eyJ...'
        )

    print(f'{args.year}年{args.month}月のデータを取得中...')
    records = fetch_records(url, key, args.year, args.month)
    print(f'  {len(records)} 件取得')

    sheet_name = f'{args.month}月'
    wb = openpyxl.load_workbook(args.template)
    if sheet_name not in wb.sheetnames:
        sys.exit(
            f'エラー: シート "{sheet_name}" が見つかりません。\n'
            f'利用可能なシート: {wb.sheetnames}'
        )
    ws = wb[sheet_name]

    for section in SECTIONS:
        daily = build_daily_totals(records, section['house_id'])
        write_section(ws, section, daily, args.year, args.month)
        print(f'  {section["name"]} セクション 書き込み完了')

    wb.save(args.output)
    print(f'保存完了: {args.output}')


if __name__ == '__main__':
    main()
